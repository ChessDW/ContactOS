from flask import Flask, render_template, request, redirect, url_for
import openpyxl, os

def inicializar_excel():
    if not os.path.exists("contactos.xlsx"):
        workbook = openpyxl.Workbook()
        hoja = workbook.active
        hoja.title = "Contactos"

        encabezados = [
            "ID",
            "Nombre",
            "Apellido",
            "Telefono",
            "Correo",
            "Descripcion",
            "Direccion",
            "Categoria",
            "Favorito"
        ]

        hoja.append(encabezados)

        workbook.save("contactos.xlsx")

def cargar_contactos():
    w = openpyxl.load_workbook("contactos.xlsx")
    s = w.active

    contactos = []

    for fila in s.iter_rows(min_row=2, values_only=True):
        contacto = {
            "id": fila[0],
            "nombre": fila[1],
            "apellido": fila[2],
            "telefono": fila[3],
            "correo": fila[4],
            "descripcion": fila[5],
            "direccion": fila[6],
            "categoria": fila[7],
            "favorito": fila[8]
        }

        contactos.append(contacto)

    return contactos
    

app = Flask(__name__)

@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form["username"]
        password = request.form["password"]

        if usuario == "admin" and password == "1234":
            return redirect(url_for("dashboard"))
        else:
            return render_template(
                "login.html",
                error="Usuario o contraseña incorrectos"
            )
        
    return render_template("login.html")

@app.route("/dashboard")
def dashboard():

    contactos = cargar_contactos()

    texto_busqueda = request.args.get("buscar", "").lower()

    solo_favoritos = request.args.get("favoritos")

    if texto_busqueda:

        contactos = [
            c for c in contactos
            if (
                texto_busqueda in c["nombre"].lower()
                or texto_busqueda in c["apellido"].lower()
                or texto_busqueda in c["telefono"].lower()
                or texto_busqueda in c["correo"].lower()
            )
        ]

    if solo_favoritos:

        contactos = [
            c for c in contactos
            if c["favorito"]
        ]

    return render_template(
        "dashboard.html",
        contactos=contactos,
        filtro="todos"
    )

@app.route("/favoritos")
def favoritos():

    contatos = cargar_contactos()

    texto_busqueda = request.args.get("buscar", "").lower()

    favoritos = [
        c for c in contatos
        if c["favorito"]
    ]

    if texto_busqueda:

        favoritos = [
            c for c in favoritos
            if (
                texto_busqueda in c["nombre"].lower()
                or texto_busqueda in c["apellido"].lower()
                or texto_busqueda in c["telefono"].lower()
                or texto_busqueda in c["correo"].lower()
            )
        ]

    return render_template(
        "dashboard.html",
        contactos=favoritos,
        filtro="favoritos"
    )

@app.route("/contacto/<int:contacto_id>")
def ver_contacto(contacto_id):

    contactos = cargar_contactos()

    for contacto in contactos:

        if int(contacto["id"]) == int(contacto_id):

            return render_template(
                "detail.html",
                contacto=contacto
            )

    return "Contacto no encontrado"



@app.route("/editar/<int:contacto_id>")
def editar_contacto(contacto_id):
    contactos = cargar_contactos()
    for contacto in contactos:
        if contacto["id"] == contacto_id:
            return render_template(
                "edit_contact.html",
                contacto=contacto
            )
        
    return "Contacto no encontrado"

@app.route("/eliminar/<int:contacto_id>")
def eliminar_contacto(contacto_id):
    workbook = openpyxl.load_workbook("contactos.xlsx")
    hoja = workbook.active
    filas = list(hoja.iter_rows(values_only=True))
    hoja.delete_rows(2, hoja.max_row)
    for fila in filas[1: ]:
        if fila[0] != contacto_id:
            hoja.append(fila)

    workbook.save("contactos.xlsx")
    return redirect(url_for("dashboard"))

@app.route("/nuevo")
def nuevo_contacto():
    return render_template("add_contact.html")

@app.route("/agregar", methods=["POST"])
def agregar_contacto():
    workbook = openpyxl.load_workbook("contactos.xlsx")
    hoja = workbook.active

    contactos = cargar_contactos()

    if contactos:
        nuevo_id = max(c["id"] for c in contactos) + 1
    else:
        nuevo_id = 1
    nombre = request.form["nombre"]
    apellido = request.form["apellido"]
    telefono = request.form["telefono"]    
    correo = request.form["correo"]
    descripcion = request.form["descripcion"]
    direccion = request.form["direccion"]
    categoria = request.form["categoria"]
    favorito = "favorito" in request.form
    if len(telefono) != 8:
        return """
    <script>
        alert('El teléfono debe tener 8 dígitos');
        window.history.back();
    </script>
    """

    if "@" not in correo:
        return """
    <script>
        alert('El teléfono debe tener 8 dígitos');
        window.history.back();
    </script>
    """

    hoja.append([
        nuevo_id,
        nombre,
        apellido,
        telefono,
        correo,
        descripcion,
        direccion,
        categoria,
        favorito
    ])

    workbook.save("contactos.xlsx")

    return redirect(url_for("dashboard"))

@app.route("/actualizar/<int:id>", methods=["POST"])
def actualizar_contacto(id):

    workbook = openpyxl.load_workbook("contactos.xlsx")
    hoja = workbook.active

    for fila in hoja.iter_rows(min_row=2):

        if fila[0].value == id:

            fila[1].value = request.form["nombre"]
            fila[2].value = request.form["apellido"]
            fila[3].value = request.form["telefono"]
            fila[4].value = request.form["correo"]
            fila[5].value = request.form["descripcion"]
            fila[6].value = request.form["direccion"]
            fila[7].value = request.form["categoria"]
            fila[8].value = "favorito" in request.form

            if len(fila[3].value) != 8:
                return """
    <script>
        alert('El teléfono debe tener 8 dígitos');
        window.history.back();
    </script>
    """

            if "@" not in fila[4].value:
                return """
    <script>
        alert('El teléfono debe tener 8 dígitos');
        window.history.back();
    </script>
    """

            break

    workbook.save("contactos.xlsx")

    return redirect(url_for("dashboard"))

@app.route("/report")
def report():

    contactos = cargar_contactos()

    total_contacts = len(contactos)

    favorite_contacts = len([
        c for c in contactos
        if c["favorito"]
    ])

    categorias = {}

    for c in contactos:
        categoria = c["categoria"]

        if categoria:
            categorias[categoria] = categorias.get(categoria, 0) + 1

    most_used_category = (
        max(categorias, key=categorias.get)
        if categorias
        else "Sin categoría"
    )

    latest_contact = contactos[-1] if contactos else None

    print(contactos)

    return render_template(
        "report.html",
        contactos=contactos,
        total_contacts=total_contacts,
        favorite_contacts=favorite_contacts,
        most_used_category=most_used_category,
        latest_contact=latest_contact
    )

@app.route("/logout")
def logout():
    return redirect(url_for("login"))

if __name__ == "__main__":
    inicializar_excel()
    app.run(debug=True)